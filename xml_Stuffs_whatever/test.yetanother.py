import openpyxl
from xml.dom import minidom
import xml.etree.ElementTree as ET
#from xml.etree.ElementTree import Element, SubElement, Comment, tostring

#%% functions
def prettify(elem):
    """Return a pretty-printed XML string for the Element.
    """
    rough_string = ET.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

def create_variable(name, value):
    globals()[name] = value
    
def setvar_folder(fold,var,value):
    fold.set(var,value)

def create_set(fold,item,el_para,para_loc):
    selset = ET.SubElement(fold, 'selectionset')
    selset.set('name', item[0])
    selset.set('guid', '')
    
    findspec = ET.SubElement(selset, 'findspec')
    findspec.set('mode', 'all')
    findspec.set('disjoint', '0')
    
    conds = ET.SubElement(findspec, 'conditions')
    
    #cond_1 - location
    cond_a = ET.SubElement(conds, 'condition')
    cond_a.set('test', 'equals')
    cond_a.set('flags', '10')
    
    cat = ET.SubElement(cond_a, 'category')
    
    name_int = ET.SubElement(cat, 'name')
    name_int.set('internal', para_loc)
    name_int.text = el_para
    
    porp = ET.SubElement(cond_a, 'property')
    
    name_int = ET.SubElement(porp, 'name')
    name_int.set('internal', 'lcldrevit_parameter_sequence_Location_PG_TEXT')
    name_int.text = 'sequence_Location'
    
    val = ET.SubElement(cond_a, 'value')
    
    name_int = ET.SubElement(val, 'data')
    name_int.set('type', 'wstring')
    name_int.text = item[2]
    
    #cond_2 - zone
    cond_b = ET.SubElement(conds, 'condition')
    cond_b.set('test', 'equals')
    cond_b.set('flags', '10')
    
    cat = ET.SubElement(cond_b, 'category')
    
    name_int = ET.SubElement(cat, 'name')
    name_int.set('internal', para_loc)
    name_int.text = el_para
    
    porp = ET.SubElement(cond_b, 'property')
    
    name_int = ET.SubElement(porp, 'name')
    name_int.set('internal', 'lcldrevit_parameter_sequence_Zone_PG_TEXT')
    name_int.text = 'sequence_Zone'
    
    val = ET.SubElement(cond_b, 'value')
    
    name_int = ET.SubElement(val, 'data')
    name_int.set('type', 'wstring')
    name_int.text = item[3]
    
    #cond_3 - task
    cond_b = ET.SubElement(conds, 'condition')
    cond_b.set('test', 'equals')
    cond_b.set('flags', '10')
    
    cat = ET.SubElement(cond_b, 'category')
    
    name_int = ET.SubElement(cat, 'name')
    name_int.set('internal', para_loc)
    name_int.text = el_para
    
    porp = ET.SubElement(cond_b, 'property')
    
    name_int = ET.SubElement(porp, 'name')
    name_int.set('internal', 'lcldrevit_parameter_sequence_Work_PG_TEXT')
    name_int.text = 'sequence_Work'
    
    val = ET.SubElement(cond_b, 'value')
    
    name_int = ET.SubElement(val, 'data')
    name_int.set('type', 'wstring')
    name_int.text = item[4]
    
    #cond_4 - part
    cond_c = ET.SubElement(conds, 'condition')
    cond_c.set('test', 'equals')
    cond_c.set('flags', '10')
    
    cat = ET.SubElement(cond_c, 'category')
    
    name_int = ET.SubElement(cat, 'name')
    name_int.set('internal', para_loc)
    name_int.text = el_para
    
    porp = ET.SubElement(cond_c, 'property')
    
    name_int = ET.SubElement(porp, 'name')
    name_int.set('internal', 'lcldrevit_parameter_sequence_WorkPart_PG_TEXT')
    name_int.text = 'sequence_WorkPart'
    
    val = ET.SubElement(cond_c, 'value')
    
    name_int = ET.SubElement(val, 'data')
    name_int.set('type', 'wstring')
    name_int.text = str(item[5])

#%% parameter

#Revit parameter
address_a = 'http://www.w3.org/2001/XMLSchema-instance'
address_b = 'http://download.autodesk.com/us/navisworks/schemas/nw-exchange-12.0.xsd'
el_para = 'Custom'
para_loc = 'LcRevitData_Custom'

#Test bed
# =============================================================================
# search_list = [['MEP + FINISHING WORK PACKAGE','R',None,None,None,None],
#                ['1 FL. Finishing work',0,None,None,None,None],
#                ['Topping Concrete 50 mm',1,None,None,None,None],
#                ['INT_F1_ALL_TOPPING_1','s','INT_F1','ALL','TOPPING','1'],
#                ['INT_F1_ALL_TOPPING_2','s','INT_F1','ALL','TOPPING','2'],
#                ['INT_F1_ALL_TOPPING_3','s','INT_F1','ALL','TOPPING','3'],
#                ['Inside CB.Wall + Plastering (Exterior wall)',1,None,None,None,None],
#                ['INT_F1_ALL_CBW_1','s','INT_F1','ALL','CBW','1'],
#                ['INT_F1_ALL_CBW_2','s','INT_F1','ALL','CBW','2'],
#                ['INT_F1_ALL_CBW_3','s','INT_F1','ALL','CBW','3'],
#                ['Zone 1 Office Area',0,None,None,None,None],
#                [' 1.1 Toilet + WC101',1,None,None,None,None],
#                ['Concrete Block wall + Plastering',2,None,None,None,None],
#                ['INT_F1_Z1.1_CBW_1','s','INT_F1','Z1.1','CBW','1'],
#                ['INT_F1_Z1.1_CBW_2','s','INT_F1','Z1.1','CBW','2'],
#                ['INT_F1_Z1.1_CBW_3','s','INT_F1','Z1.1','CBW','3'],]
# =============================================================================

#import excel para, Interior
path = "C:\\Users\\d.suniwat\\Documents\\Work\\07-Method\\2024-11-27-Interior\\Doc_out\\task_phase.xlsx"
wb_1 = openpyxl.load_workbook(path, data_only=True)
ws_1 = wb_1["ForNVM"]

#save path
save_path_file = "C:\\Users\\d.suniwat\\Documents\\Work\\07-Method\\2024-11-27-Interior\\Doc_out\\task_phase.0113.xml"  

#import excel para, Landscape
# =============================================================================
# path = "C:\\Users\\d.suniwat\\Documents\\Work\\07-Method\\2024-11-20-Landscape\\Doc_in\\MasterPlan\\Breakdownschedule\\LZ_Name.xlsx"
# wb_1 = openpyxl.load_workbook(path, data_only=True)
# 
# # =============================================================================
# # ws_1 = wb_1["NVMSet"] #LA
# # #save path LA
# # save_path_file = "C:\\Users\\d.suniwat\\Documents\\Work\\07-Method\\2024-11-20-Landscape\\Doc_in\\MasterPlan\\Breakdownschedule\\task.0127.xml" 
# # =============================================================================
# 
# # =============================================================================
# # ws_1 = wb_1["NVMSet.VH"] #LA Vehicle
# # #save path LA Vehicle 
# # save_path_file = "C:\\Users\\d.suniwat\\Documents\\Work\\07-Method\\2024-11-20-Landscape\\Doc_in\\MasterPlan\\Breakdownschedule\\task.0127.vh.xml"  
# =============================================================================
# =============================================================================

#%% main

#import list from excel
search_list = []
for i, row in enumerate(ws_1.iter_rows()):
    #if i > 0 and i < 20:
    #exclude header
    if i > 0:
        r1 = []
        for cell in row:
            r1.append(cell.value)
            
        #print(r1)            
        search_list.append(r1)

#print(search_list)

#xml headers
head = ET.Element('exchange')
head.set('xmlns:xsi', address_a)
head.set('xsi:noNamespaceSchemaLocation', address_b)
head.set('units', 'ft')
head.set('filename', '')
head.set('filepath', '')

selsets = ET.SubElement(head, 'selectionsets')

#folders template
fld_a =  ET.SubElement(selsets, 'viewfolder')
setvar_folder(fld_a, 'guid', '')
setvar_folder(fld_a, 'name', 'Empty')

# =============================================================================
# 
# fld_b =  ET.SubElement(fld_a, 'viewfolder')
# setvar_folder(fld_b, 'guid', '')
# setvar_folder(fld_b, 'name', 'Empty')
# 
# fld_c =  ET.SubElement(fld_b, 'viewfolder')
# setvar_folder(fld_c, 'guid', '')
# setvar_folder(fld_c, 'name', 'Empty')
# 
# fld_d =  ET.SubElement(fld_c, 'viewfolder')
# setvar_folder(fld_d, 'guid', '')
# setvar_folder(fld_d, 'name', 'Empty')
# =============================================================================

# =============================================================================
# fld_e =  ET.SubElement(fld_d, 'viewfolder')
# setvar_folder(fld_e, 'guid', '')
# setvar_folder(fld_a, 'name', 'Empty')
# =============================================================================

# =============================================================================
# group_folder = [fld_a,
#                 fld_b,
#                 fld_c,
#                 fld_d]
# =============================================================================

# =============================================================================
# group_folder = [fld_a,
#                 fld_b,
#                 fld_c,
#                 fld_d,
#                 fld_e]
# =============================================================================

#create folders and sets
index_f = 0
for item in search_list:
    if item[1] == 'R':
        setvar_folder(fld_a, 'name', item[0])
        index_f = 0
    elif item[1] == 0:
        fld_b =  ET.SubElement(fld_a, 'viewfolder')
        setvar_folder(fld_b, 'guid', '')
        setvar_folder(fld_b, 'name', item[0])
        index_f = 1
    elif item[1] == 1:
        fld_c =  ET.SubElement(fld_b, 'viewfolder')
        setvar_folder(fld_c, 'guid', '')
        setvar_folder(fld_c, 'name', item[0])
        index_f = 2
    elif item[1] == 2:
        fld_d =  ET.SubElement(fld_c, 'viewfolder')
        setvar_folder(fld_d, 'guid', '')
        setvar_folder(fld_d, 'name', item[0])
        index_f = 3
    elif item[1] == 3:
        fld_e =  ET.SubElement(fld_d, 'viewfolder')
        setvar_folder(fld_e, 'guid', '')
        setvar_folder(fld_e, 'name', item[0])
        index_f = 4
    elif item[1] == 's':
        #print(item)
        match index_f :
            case 0:
                create_set(fld_a,item,el_para,para_loc)            
            case 1:
                create_set(fld_b,item,el_para,para_loc)
            case 2:
                create_set(fld_c,item,el_para,para_loc)            
            case 3:
                create_set(fld_d,item,el_para,para_loc)
# =============================================================================
#         print(group_folder[index_f].tag)
#         print(group_folder[index_f].attrib)
# =============================================================================
    
xml_str = prettify(head)
# =============================================================================
# print(xml_str)
# =============================================================================

#save file
xml_str = prettify(head)
with open(save_path_file, "w") as f: 
    f.write(xml_str)  

